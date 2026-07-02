import json
from pathlib import Path

import openpyxl


SOURCE = Path("/Users/haroon/Downloads/Digital_Marketing_Fit_Salary_Model.xlsx")
OUT = Path(__file__).resolve().parents[1] / "src" / "data" / "workbookData.json"


def clean(value):
    if value is None:
        return None
    if isinstance(value, str):
        return value.strip()
    return value


def rows(ws_name):
    wb = openpyxl.load_workbook(SOURCE, data_only=False)
    ws = wb[ws_name]
    headers = [clean(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
    data = []
    for r in range(2, ws.max_row + 1):
        item = {}
        for c, header in enumerate(headers, 1):
            if not header:
                continue
            value = clean(ws.cell(r, c).value)
            if value is not None:
                item[header] = value
        if item:
            data.append(item)
    return data


def survey_questions():
    wb = openpyxl.load_workbook(SOURCE, data_only=False)
    ws = wb["Survey Questions"]
    questions = []
    current = None
    for r in range(1, ws.max_row + 1):
        number = clean(ws.cell(r, 1).value)
        question = clean(ws.cell(r, 2).value)
        option = clean(ws.cell(r, 3).value)
        field_option = clean(ws.cell(r, 7).value)
        if isinstance(number, int) or (isinstance(number, str) and number.isdigit()):
            current = {
                "number": int(number),
                "question": question,
                "inputType": option,
                "options": [],
            }
            questions.append(current)
        elif current and option:
            current["options"].append(option)
        if field_option:
            # Column G is the field-of-study option list.
            pass
    return questions


def field_options():
    wb = openpyxl.load_workbook(SOURCE, data_only=False)
    ws = wb["Survey Questions"]
    values = []
    for r in range(1, ws.max_row + 1):
        value = clean(ws.cell(r, 7).value)
        if value:
            values.append(value)
    return values


def main():
    payload = {
        "salaryLookup": rows("Salary Lookup"),
        "fieldFitScores": rows("Field Fit Scores"),
        "answerScoreRules": rows("Answer Score Rules"),
        "salaryFactors": rows("Salary Factors"),
        "roleGroups": rows("Role Groups"),
        "surveyQuestions": survey_questions(),
        "fieldOptions": field_options(),
    }
    OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
